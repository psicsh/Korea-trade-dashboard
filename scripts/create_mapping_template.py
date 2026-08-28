#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "mti_hsk_mapping.xlsx"


def main() -> int:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    rows = [
        ["한국무역 대시보드 2026 MTI-HSK 연계표 자리표시자"],
        ["이 파일에는 실제 HSK-MTI 매핑이 포함되어 있지 않습니다."],
        ["한국무역협회가 공개한 공식 2026 연계표로 이 파일을 교체한 뒤 최초 구축을 실행하세요."],
        ["로더는 HSK/HSK10/HS코드/세번 열과 MTI/MTI6/MTI코드 열을 자동 인식합니다."],
    ]
    for row in rows:
        readme.append(row)
    readme["A1"].font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 105

    mapping = workbook.create_sheet("mapping")
    mapping.append(["HSK10", "HSK_NAME_KO", "MTI6", "MTI_NAME_KO"])
    for cell in mapping[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column, width in {"A": 16, "B": 48, "C": 14, "D": 36}.items():
        mapping.column_dimensions[column].width = width
    mapping.freeze_panes = "A2"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"created: {OUTPUT.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
